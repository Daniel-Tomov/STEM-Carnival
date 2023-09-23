from flask import Flask, jsonify, render_template, request, make_response
from flask_compress import Compress
import pandas as pd
from datetime import datetime
import xlsxwriter
from openpyxl import load_workbook

compress = Compress()
def start_app():
    app = Flask(__name__)
    compress.init_app(app)
    return app
app = Flask(__name__)
app.jinja_env.trim_blocks = True
app.jinja_env.lstrip_blocks = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config.update(
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Strict',
)

# example structure of one of the keys and values
# uuid: {current_game_uuid: None, data: {<uuids of sensors and their data>}}
currently_running_games = {}

def get_player_uuid_from_game_uuid(game_uuid):
  for player_uuid in currently_running_games:
    if currently_running_games[player_uuid]["current_game_uuid"] == game_uuid:
      return player_uuid
  # will return None if no one is playing the game
  return None

def get_time(): return datetime.now().strftime("%H:%M:%S")


@app.route('/', methods=["POST", "GET"])
def index():
  return make_response(render_template('index.html', currently_running_games=currently_running_games))


@app.route('/game/<string:game_uuid>', methods=["POST", 'GET']) 
def game(game_uuid):
  # get the player UUID from the request
  player_uuid = request.data.decode('utf-8')

  # check if the player is playing another game
  # this prevents people sharing a card
  if player_uuid in currently_running_games.keys():
    
    # check if the player which scanned the card is playing the game they scanned at
    if currently_running_games[player_uuid]["current_game_uuid"] == game_uuid:
      #
      currently_running_games.pop(player_uuid)
      return jsonify({"status":"ending"})
    # if the
    else:
      return jsonify({"status":"player in use"})
  
  # player uuid is not currently playing a game
  else:
    # there is not a player currently playing the game
    if get_player_uuid_from_game_uuid(game_uuid) == None:
      currently_running_games[player_uuid] = {"current_game_uuid": game_uuid}
      return jsonify({"status": "starting"})
    
    # there is a player currently playing the game
    else:
      return jsonify({"status":"game in use"})

  return jsonify({"status":"error"})

@app.route('/sensor/<string:game_uuid>/<string:sensor_name>', methods=["POST", 'GET']) 
def sensor(game_uuid, sensor_name):
  # get the sensor data
  data = request.data.decode('utf-8')
  # get the player uuid
  player_uuid = get_player_uuid_from_game_uuid(game_uuid)
  # checks if the player is playing a game
  if player_uuid == None:
    return jsonify({"status":"there is not a player playing that game right now"})
  # put the data from the sensor in an array to save it
  save_data = [get_time(), sensor_name, data]

  # save the data to the "player_data" folder and the workbook name with the player uuid
  # each worksheet has the name of the game uuid
  # each sheet contains the time, sensor name, and value from the sensor
  try:
    with pd.ExcelWriter("player_data/" + player_uuid + ".xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
      pd.DataFrame([save_data], columns=["time", "sensor_name", "sensor_value"]).to_excel(writer, header=False, startrow=writer.sheets[game_uuid].max_row, index=False, sheet_name=game_uuid)

  # if the player is playing for the first time, they won't have a workbook associated to their uuid
  except FileNotFoundError:
    xlsx_File = xlsxwriter.Workbook('player_data/' + player_uuid + ".xlsx")
    xlsx_File.add_worksheet(game_uuid)
    xlsx_File.close()

    with pd.ExcelWriter("player_data/" + player_uuid + ".xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
      pd.DataFrame([save_data], columns=["time", "sensor_name", "sensor_value"]).to_excel(writer, header=False, startrow=writer.sheets[game_uuid].max_row, index=False, sheet_name=game_uuid)
  # if the player has not played the game before, they won't have a worksheet associated to them for that game
  except KeyError:
    book = load_workbook('player_data/' + player_uuid + '.xlsx')
    new_sheet = book.create_sheet(game_uuid)
    book.save('player_data/' + player_uuid + '.xlsx')

    with pd.ExcelWriter("player_data/" + player_uuid + ".xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
      pd.DataFrame([save_data], columns=["time", "sensor_name", "sensor_value"]).to_excel(writer, header=False, startrow=writer.sheets[game_uuid].max_row, index=False, sheet_name=game_uuid)

  # return something
  return jsonify({"status":"success"})

@app.errorhandler(404)
def not_found(e):
  return ""


if __name__ == "__main__":
  app.run(host="0.0.0.0", port=5555, debug=True, use_reloader=True)